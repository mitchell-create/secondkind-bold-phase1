"""XLSX writer for the creative matrix — produces the human-facing workbook.

One workbook with four sheets matching the four matrix tabs. Each sheet has
a header row, frozen top row, sensible column widths, and wrap-text on long
cells. Common columns (format_fit, static_treatment, psychology_trigger,
trending_*) trail the tab-specific A-D columns so a glance at the left side
reads strategically and a glance at the right side reads tactically.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.matrix import CreativeMatrix, MatrixRow, MatrixTab

# ─── Column specs per tab ───────────────────────────────────────────────────

# Each entry: (header_label, attr_or_callable, width_chars)
# `attr_or_callable` is either a string attribute name on MatrixRow OR a
# callable `(row) -> str` for computed cells (e.g. flattening trending objects).


def _join_list(values: list[str]) -> str:
    return ", ".join(values) if values else ""


def _trending_match_concept(row: MatrixRow) -> str:
    return row.trending_match.concept if row.trending_match else ""


def _trending_match_execution(row: MatrixRow) -> str:
    return row.trending_match.execution if row.trending_match else ""


def _trending_match_format_id(row: MatrixRow) -> str:
    return row.trending_match.format_id if row.trending_match else ""


def _trending_next_idea(row: MatrixRow) -> str:
    return row.trending_next.idea if row.trending_next else ""


def _trending_next_execution(row: MatrixRow) -> str:
    return row.trending_next.execution if row.trending_next else ""


_BRIDGE_COLUMNS: list[tuple[str, Any, int]] = [
    ("Format fit", lambda r: _join_list(r.format_fit), 28),
    ("Static treatment", "static_treatment", 60),
    ("Psychology trigger", "psychology_trigger", 20),
    ("Trending match — concept", _trending_match_concept, 28),
    ("Trending match — execution", _trending_match_execution, 60),
    ("Trending match — format_id", _trending_match_format_id, 22),
    ("Trending next — idea", _trending_next_idea, 28),
    ("Trending next — execution", _trending_next_execution, 60),
]


TAB_COLUMNS: dict[MatrixTab, list[tuple[str, Any, int]]] = {
    MatrixTab.PAIN_VS_COMPETITOR: [
        ("ID", "id", 12),
        ("Complaint", "complaint", 40),
        ("Source", lambda r: _join_list(r.source_type), 22),
        ("What people say", "what_people_say", 50),
        ("Our advantage", "our_advantage", 40),
        ("Positioning angle", "positioning_angle", 50),
        *_BRIDGE_COLUMNS,
    ],
    MatrixTab.WHAT_THEY_LOVE: [
        ("ID", "id", 12),
        ("Emotional trigger", "emotional_trigger", 40),
        ("Source", lambda r: _join_list(r.source_type), 22),
        ("What people say", "what_people_say", 50),
        ("Hook / angle to amplify", "hook_or_angle_to_amplify", 50),
        *_BRIDGE_COLUMNS,
    ],
    MatrixTab.WISHES_GAPS: [
        ("ID", "id", 12),
        ("What they wish existed", "wished_product", 40),
        ("Source", lambda r: _join_list(r.source_type), 22),
        ("Real frustration", "real_frustration", 40),
        ("Script / static opportunity", "script_or_static_opportunity", 50),
        *_BRIDGE_COLUMNS,
    ],
    MatrixTab.HOOK_ANGLES: [
        ("ID", "id", 12),
        ("Hook angle", "hook_angle", 50),
        ("Tactic", "tactic", 26),
        ("Story / setup", "story_setup", 50),
        ("Why it works", "why_it_works", 50),
        *_BRIDGE_COLUMNS,
    ],
}


_TAB_TITLE: dict[MatrixTab, str] = {
    MatrixTab.PAIN_VS_COMPETITOR: "Pain vs Competitor",
    MatrixTab.WHAT_THEY_LOVE: "What They Love",
    MatrixTab.WISHES_GAPS: "Wishes & Gaps",
    MatrixTab.HOOK_ANGLES: "Hook Angles",
}


# ─── Styling ────────────────────────────────────────────────────────────────


def _header_style() -> tuple[Font, PatternFill, Alignment, Border]:
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="374151")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, align, border


def _cell_style() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


# ─── Writer ─────────────────────────────────────────────────────────────────


def _resolve_value(row: MatrixRow, accessor: Any) -> str:
    if callable(accessor):
        return str(accessor(row) or "")
    value = getattr(row, accessor, "")
    if value is None:
        return ""
    if isinstance(value, list):
        return _join_list(value)
    return str(value)


def _write_tab(ws, tab: MatrixTab, rows: list[MatrixRow]) -> None:
    columns = TAB_COLUMNS[tab]
    h_font, h_fill, h_align, h_border = _header_style()
    body_align = _cell_style()

    # Header row
    for col_idx, (label, _accessor, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = h_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Body rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, accessor, _width) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_resolve_value(row, accessor))
            cell.alignment = body_align

    # Freeze header + raise default row height for readability
    ws.freeze_panes = "A2"
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 75


def write_matrix_xlsx(matrix: CreativeMatrix, out_path: Path) -> Path:
    """Write the matrix to an XLSX workbook with one sheet per tab."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet — we'll add ours in order.
    default = wb.active
    wb.remove(default)

    tab_order = [
        MatrixTab.PAIN_VS_COMPETITOR,
        MatrixTab.WHAT_THEY_LOVE,
        MatrixTab.WISHES_GAPS,
        MatrixTab.HOOK_ANGLES,
    ]
    for tab in tab_order:
        rows = getattr(matrix, tab.value)
        ws = wb.create_sheet(title=_TAB_TITLE[tab])
        _write_tab(ws, tab, rows)

    wb.save(out_path)
    return out_path
